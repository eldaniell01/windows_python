import typing
from openpyxl import load_workbook
from PyQt6 import QtCore, uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QTableWidgetItem, QFileDialog

class ExcelDoc(QMainWindow):
    def __init__(self):
        super().__init__()
        self.excel = uic.loadUi("interfaz/Excel.ui")
        self.excel.show()
        self.excel.btnexcel.clicked.connect(self.pathExcel)
        self.excel.return1.clicked.connect(self.regreso)
        
    def pathExcel(self):
        folder = QFileDialog()
        folder_path, __= folder.getOpenFileName(None, 'ABRIR ARCHIVO', '', 'xlsx (*.xlsx)')
        self.excel.textpath.setText(folder_path)
        self.loadExcel(folder_path)
        
    def loadExcel(self, path):
        workbook = load_workbook(filename=path)
        sheet = workbook.active
        colums = ["NOMBRE", "APELLIDO", "EDAD"]
        self.excel.table.setColumnCount(len(colums))
        for row in sheet.iter_rows(values_only= True):
            row_index = self.excel.table.rowCount()
            self.excel.table.insertRow(row_index)
            for col_index, value in enumerate(row):
                self.excel.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
        self.excel.table.setHorizontalHeaderLabels(colums)  
        
    def regreso(self):
        from interfaz.principal import Windows
        self.regreso = Windows()
        self.excel.close()